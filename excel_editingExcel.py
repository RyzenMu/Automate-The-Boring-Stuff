import openpyxl
import os

os.chdir("C:\\users\\x\\Desktop")
workbook = openpyxl.load_workbook('example1.xlsx')
sheet = workbook['Sheet1']
print(sheet.cell(column=1, row=1).value)
sheet['E5'] = 'Mongo'
workbook.save('example2.xlsx')
newSheet = workbook.create_sheet(index=0, title='Sheet12')
newSheet.title = "Names"
print(workbook.get_sheet_names())
workbook.save('example3.xlsx')

#Creating a sheet
wb = openpyxl.Workbook()
sheet = wb.get_sheet_names()
print(sheet[0])
# print(sheet[0].cell(column=1, row=1).value == None) 
