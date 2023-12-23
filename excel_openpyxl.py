import openpyxl
import os

os.chdir("C:\\Users\\x\\Desktop")

workbook = openpyxl.load_workbook('example1.xlsx')
print(type(workbook))

Sheet1 = workbook['Sheet1']
print(type(Sheet1))

sheetNames = workbook.get_sheet_names()
print(sheetNames)

# to get the value in the cell use squarebracket []
cellA4 = Sheet1['A4']
print(cellA4)
print(cellA4.value)

print(Sheet1['c3'].value)

print(Sheet1.cell(column=5, row=3))

# iterating using for loop
for i in range (1, 10):
    print(i, Sheet1['B'+str(i)].value, Sheet1.cell(row=i, column=3).value)